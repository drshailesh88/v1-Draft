"""
Data Extraction API - Agent D
Extract tables, figures, and data from PDF documents.

Features:
- Table extraction from PDFs using PDFPlumber
- Figure detection and caption extraction
- CSV/Excel export functionality
- Metadata preservation (page number, title)
"""

from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime
import tempfile
import os
import io
import csv
import re
import base64
import uuid

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font

from core.database import supabase

router = APIRouter()


# ============================================================================
# Pydantic Models
# ============================================================================

class TableData(BaseModel):
    """Represents a single extracted table"""
    table_id: int
    page_number: int
    title: Optional[str] = None
    headers: List[str] = Field(default_factory=list)
    rows: List[List[str]] = Field(default_factory=list)
    row_count: int = 0
    column_count: int = 0


class FigureData(BaseModel):
    """Represents an extracted figure/image"""
    figure_id: int
    page_number: int
    caption: Optional[str] = None
    bbox: Optional[Dict[str, float]] = None  # x0, y0, x1, y1
    width: Optional[float] = None
    height: Optional[float] = None
    image_data: Optional[str] = None  # Base64 encoded image


class ExtractionMetadata(BaseModel):
    """Metadata about the extraction process"""
    filename: str
    total_pages: int
    extraction_timestamp: str
    tables_found: int
    figures_found: int


class ExtractTablesResponse(BaseModel):
    """Response for table extraction"""
    success: bool
    metadata: ExtractionMetadata
    tables: List[TableData]


class ExtractFiguresResponse(BaseModel):
    """Response for figure extraction"""
    success: bool
    metadata: ExtractionMetadata
    figures: List[FigureData]


class FullExtractionResponse(BaseModel):
    """Response for full document extraction"""
    success: bool
    metadata: ExtractionMetadata
    tables: List[TableData]
    figures: List[FigureData]


class ExportRequest(BaseModel):
    """Request for exporting data"""
    tables: List[TableData]
    filename: Optional[str] = "extracted_data"


class SaveExtractionRequest(BaseModel):
    """Request to save extraction to database"""
    document_id: Optional[str] = None
    user_id: Optional[str] = None
    extraction_type: str  # 'tables', 'figures', or 'full'
    data: Dict[str, Any]


class SaveExtractionResponse(BaseModel):
    """Response from saving extraction"""
    success: bool
    extraction_id: str
    message: str


class ExtractAndSaveResponse(BaseModel):
    """Response for extract and save combined operation"""
    success: bool
    extraction_id: str
    metadata: ExtractionMetadata
    tables: List[TableData]
    figures: List[FigureData]
    message: str


# ============================================================================
# Helper Functions
# ============================================================================

def clean_cell_value(value: Any) -> str:
    """Clean and normalize cell values from PDFPlumber"""
    if value is None:
        return ""
    # Convert to string and clean whitespace
    text = str(value).strip()
    # Replace multiple whitespace with single space
    text = re.sub(r'\s+', ' ', text)
    return text


def extract_table_title(table_bbox: tuple, page_text: str) -> Optional[str]:
    """
    Try to extract a title for the table by looking at text above it.
    Looks for common patterns like "Table 1:", "Table 1.", etc.
    """
    if not table_bbox:
        return None

    # Search for table title patterns in the page text
    try:
        patterns = [
            r'(Table\s+\d+[.:]\s*[^\n]+)',
            r'(TABLE\s+\d+[.:]\s*[^\n]+)',
            r'(Tab\.\s*\d+[.:]\s*[^\n]+)',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, page_text, re.IGNORECASE)
            if matches:
                # Return the first match as potential title
                return matches[0].strip()
    except Exception:
        pass

    return None


def extract_figure_captions(page_text: str) -> List[Dict[str, Any]]:
    """
    Extract figure captions from page text.
    Looks for patterns like "Figure 1:", "Fig. 1.", etc.
    """
    captions = []
    seen_numbers = set()

    patterns = [
        r'(Figure\s+(\d+)[.:]\s*[^\n]+)',
        r'(FIGURE\s+(\d+)[.:]\s*[^\n]+)',
        r'(Fig\.\s*(\d+)[.:]\s*[^\n]+)',
        r'(FIG\.\s*(\d+)[.:]\s*[^\n]+)',
    ]

    for pattern in patterns:
        matches = re.findall(pattern, page_text, re.IGNORECASE)
        for match in matches:
            full_caption = match[0].strip()
            fig_num = int(match[1])

            # Avoid duplicates
            if fig_num not in seen_numbers:
                seen_numbers.add(fig_num)
                captions.append({
                    'figure_number': fig_num,
                    'caption': full_caption
                })

    # Sort by figure number
    captions.sort(key=lambda x: x['figure_number'])
    return captions


def process_pdf_tables(pdf_path: str) -> tuple[List[TableData], int]:
    """
    Extract all tables from a PDF file using PDFPlumber.
    Returns a list of TableData objects and total page count.
    """
    tables = []
    table_id = 0
    total_pages = 0

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)

        for page_num, page in enumerate(pdf.pages, start=1):
            # Get page text for title extraction
            page_text = page.extract_text() or ""

            # Extract tables from this page
            page_tables = page.extract_tables()

            # Get table bounding boxes for title extraction
            found_tables = page.find_tables()

            for idx, table in enumerate(page_tables):
                if not table or len(table) == 0:
                    continue

                table_id += 1

                # Get bounding box for this specific table
                table_bbox = found_tables[idx].bbox if idx < len(found_tables) else None

                # Try to extract title
                title = extract_table_title(table_bbox, page_text)

                # Process table data
                headers = []
                rows = []

                # First row is typically headers
                if len(table) > 0:
                    headers = [clean_cell_value(cell) for cell in table[0]]

                    # Remaining rows are data
                    for row in table[1:]:
                        cleaned_row = [clean_cell_value(cell) for cell in row]
                        rows.append(cleaned_row)

                table_data = TableData(
                    table_id=table_id,
                    page_number=page_num,
                    title=title,
                    headers=headers,
                    rows=rows,
                    row_count=len(rows),
                    column_count=len(headers)
                )
                tables.append(table_data)

    return tables, total_pages


def process_pdf_figures(pdf_path: str, include_image_data: bool = False) -> tuple[List[FigureData], int]:
    """
    Extract figures/images from a PDF file using PDFPlumber.
    Returns a list of FigureData objects and total page count.
    """
    figures = []
    figure_id = 0
    total_pages = 0

    # Track all captions across pages
    all_captions = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)

        # First pass: collect all captions
        for page_num, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text() or ""
            page_captions = extract_figure_captions(page_text)
            for cap in page_captions:
                cap['page_number'] = page_num
            all_captions.extend(page_captions)

        # Second pass: extract images and match captions
        for page_num, page in enumerate(pdf.pages, start=1):
            # Extract images from this page
            images = page.images

            for img in images:
                figure_id += 1

                # Get image bounding box
                bbox = {
                    'x0': img.get('x0', 0),
                    'y0': img.get('top', 0),
                    'x1': img.get('x1', 0),
                    'y1': img.get('bottom', 0)
                }

                width = bbox['x1'] - bbox['x0']
                height = bbox['y1'] - bbox['y0']

                # Skip very small images (likely artifacts)
                if width < 50 or height < 50:
                    figure_id -= 1  # Don't count this as a figure
                    continue

                # Try to match with a caption
                caption = None

                # Look for captions on the same page first
                page_caps = [c for c in all_captions if c['page_number'] == page_num]
                if page_caps:
                    # Find caption with matching figure number or first available
                    for cap in page_caps:
                        if cap['figure_number'] == figure_id:
                            caption = cap['caption']
                            break
                    if not caption and page_caps:
                        # Use first unmatched caption on this page
                        caption = page_caps[0]['caption']

                # Optionally extract image data
                image_data = None
                if include_image_data:
                    try:
                        # Get the image from the page
                        img_obj = page.crop((bbox['x0'], bbox['y0'], bbox['x1'], bbox['y1']))
                        img_pil = img_obj.to_image(resolution=150)

                        # Convert to base64
                        buffer = io.BytesIO()
                        img_pil.save(buffer, format='PNG')
                        image_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
                    except Exception:
                        # Image extraction can fail for various reasons
                        pass

                figure_data = FigureData(
                    figure_id=figure_id,
                    page_number=page_num,
                    caption=caption,
                    bbox=bbox,
                    width=width,
                    height=height,
                    image_data=image_data
                )
                figures.append(figure_data)

    return figures, total_pages


def generate_extraction_id() -> str:
    """Generate a unique extraction ID"""
    return str(uuid.uuid4())


# ============================================================================
# API Endpoints
# ============================================================================

@router.post("/upload-and-extract-tables", response_model=ExtractTablesResponse)
async def upload_and_extract_tables(
    file: UploadFile = File(..., description="PDF file to extract tables from")
):
    """
    Upload a PDF file and extract all tables from it.

    Returns structured table data with:
    - Page numbers
    - Table titles (if detectable)
    - Headers and data rows
    - Row and column counts
    """
    # Validate file type
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400,
            detail="Only PDF files are supported. Please upload a .pdf file."
        )

    # Create temporary file to process
    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_path = temp_file.name

        # Extract tables
        tables, total_pages = process_pdf_tables(temp_path)

        # Build response
        metadata = ExtractionMetadata(
            filename=file.filename,
            total_pages=total_pages,
            extraction_timestamp=datetime.utcnow().isoformat(),
            tables_found=len(tables),
            figures_found=0
        )

        return ExtractTablesResponse(
            success=True,
            metadata=metadata,
            tables=tables
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to extract tables: {str(e)}"
        )
    finally:
        # Clean up temp file
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


@router.post("/upload-and-extract-figures", response_model=ExtractFiguresResponse)
async def upload_and_extract_figures(
    file: UploadFile = File(..., description="PDF file to extract figures from"),
    include_image_data: bool = Query(
        default=False,
        description="Include base64-encoded image data in response"
    )
):
    """
    Upload a PDF file and extract all figures/images from it.

    Returns figure data with:
    - Page numbers
    - Captions (if detectable)
    - Bounding boxes
    - Dimensions
    - Optionally, base64-encoded image data
    """
    # Validate file type
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400,
            detail="Only PDF files are supported. Please upload a .pdf file."
        )

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_path = temp_file.name

        # Extract figures
        figures, total_pages = process_pdf_figures(temp_path, include_image_data)

        # Build response
        metadata = ExtractionMetadata(
            filename=file.filename,
            total_pages=total_pages,
            extraction_timestamp=datetime.utcnow().isoformat(),
            tables_found=0,
            figures_found=len(figures)
        )

        return ExtractFiguresResponse(
            success=True,
            metadata=metadata,
            figures=figures
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to extract figures: {str(e)}"
        )
    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


@router.post("/upload-and-extract-all", response_model=FullExtractionResponse)
async def upload_and_extract_all(
    file: UploadFile = File(..., description="PDF file to extract data from"),
    include_image_data: bool = Query(
        default=False,
        description="Include base64-encoded image data for figures"
    )
):
    """
    Upload a PDF file and extract both tables and figures.

    This is a convenience endpoint that combines table and figure extraction
    in a single request.
    """
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400,
            detail="Only PDF files are supported. Please upload a .pdf file."
        )

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_path = temp_file.name

        # Extract both tables and figures
        tables, total_pages = process_pdf_tables(temp_path)
        figures, _ = process_pdf_figures(temp_path, include_image_data)

        metadata = ExtractionMetadata(
            filename=file.filename,
            total_pages=total_pages,
            extraction_timestamp=datetime.utcnow().isoformat(),
            tables_found=len(tables),
            figures_found=len(figures)
        )

        return FullExtractionResponse(
            success=True,
            metadata=metadata,
            tables=tables,
            figures=figures
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to extract data: {str(e)}"
        )
    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


@router.post("/extract-and-save", response_model=ExtractAndSaveResponse)
async def extract_and_save(
    file: UploadFile = File(..., description="PDF file to extract data from"),
    include_image_data: bool = Query(
        default=False,
        description="Include base64-encoded image data for figures"
    ),
    document_id: Optional[str] = Query(default=None, description="Document ID to associate"),
    user_id: Optional[str] = Query(default=None, description="User ID (optional in dev mode)")
):
    """
    Upload a PDF, extract all data (tables and figures), and save to database.

    This is a convenience endpoint that combines extraction and persistence
    in a single request. Returns the extraction ID for later retrieval.
    """
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=400,
            detail="Only PDF files are supported. Please upload a .pdf file."
        )

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_path = temp_file.name

        # Extract both tables and figures
        tables, total_pages = process_pdf_tables(temp_path)
        figures, _ = process_pdf_figures(temp_path, include_image_data)

        metadata = ExtractionMetadata(
            filename=file.filename,
            total_pages=total_pages,
            extraction_timestamp=datetime.utcnow().isoformat(),
            tables_found=len(tables),
            figures_found=len(figures)
        )

        # Prepare data for database
        extraction_data = {
            "metadata": metadata.model_dump(),
            "tables": [t.model_dump() for t in tables],
            "figures": [f.model_dump() for f in figures]
        }

        # Save to database
        record = {
            "extraction_type": "full",
            "data": extraction_data,
            "created_at": datetime.utcnow().isoformat()
        }

        if document_id:
            record["document_id"] = document_id
        if user_id:
            record["user_id"] = user_id

        response = supabase.table("data_extractions").insert(record).execute()

        if not response.data:
            raise HTTPException(
                status_code=500,
                detail="Failed to save extraction to database"
            )

        extraction_id = response.data[0]["id"]

        return ExtractAndSaveResponse(
            success=True,
            extraction_id=str(extraction_id),
            metadata=metadata,
            tables=tables,
            figures=figures,
            message=f"Extracted {len(tables)} tables and {len(figures)} figures, saved to database"
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to extract and save: {str(e)}"
        )
    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


@router.post("/export-csv")
async def export_tables_to_csv(request: ExportRequest):
    """
    Export extracted tables to CSV format.

    If multiple tables are provided, they are concatenated with a separator row
    indicating the table number and page.

    Returns a downloadable CSV file.
    """
    if not request.tables:
        raise HTTPException(
            status_code=400,
            detail="No tables provided for export"
        )

    try:
        output = io.StringIO()
        writer = csv.writer(output)

        for i, table in enumerate(request.tables):
            # Add table separator/header
            if i > 0:
                writer.writerow([])  # Empty row separator

            writer.writerow([f"=== Table {table.table_id} (Page {table.page_number}) ==="])

            if table.title:
                writer.writerow([f"Title: {table.title}"])

            # Write headers
            if table.headers:
                writer.writerow(table.headers)

            # Write data rows
            for row in table.rows:
                writer.writerow(row)

        # Create streaming response
        output.seek(0)
        filename = f"{request.filename}.csv"

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to export CSV: {str(e)}"
        )


@router.post("/export-excel")
async def export_tables_to_excel(request: ExportRequest):
    """
    Export extracted tables to Excel format.

    Each table is placed in a separate sheet named by table ID and page number.
    Includes metadata in each sheet.

    Returns a downloadable .xlsx file.
    """
    if not request.tables:
        raise HTTPException(
            status_code=400,
            detail="No tables provided for export"
        )

    try:
        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Define bold font for headers
        bold_font = Font(bold=True)

        for table in request.tables:
            # Create sheet for this table
            sheet_name = f"Table_{table.table_id}_Page_{table.page_number}"
            # Excel sheet names have a 31 character limit
            sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)

            # Add metadata rows
            ws.append([f"Table ID: {table.table_id}"])
            ws.append([f"Page Number: {table.page_number}"])
            if table.title:
                ws.append([f"Title: {table.title}"])
            ws.append([f"Rows: {table.row_count}, Columns: {table.column_count}"])
            ws.append([])  # Empty row separator

            # Add headers
            if table.headers:
                ws.append(table.headers)
                # Bold the header row
                for cell in ws[ws.max_row]:
                    cell.font = bold_font

            # Add data rows
            for row in table.rows:
                ws.append(row)

            # Auto-adjust column widths
            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value or ""))
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width

        # If no tables were added, create an info sheet
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="No Tables")
            ws.append(["No tables found in the document"])

        # Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{request.filename}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to export Excel: {str(e)}"
        )


@router.post("/export-single-table-csv")
async def export_single_table_csv(
    table: TableData,
    filename: str = Query(default="table", description="Filename without extension")
):
    """
    Export a single table to CSV format.

    Cleaner output without separators, suitable for direct data use.
    """
    try:
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        if table.headers:
            writer.writerow(table.headers)

        # Write data rows
        for row in table.rows:
            writer.writerow(row)

        output.seek(0)

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}.csv"
            }
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to export CSV: {str(e)}"
        )


@router.post("/save-extraction", response_model=SaveExtractionResponse)
async def save_extraction_to_database(request: SaveExtractionRequest):
    """
    Save extraction results to the database.

    This endpoint stores the extracted data in the data_extractions table
    for later retrieval and analysis.

    Note: In development mode, user_id is optional.
    """
    try:
        # Prepare the record
        record = {
            "extraction_type": request.extraction_type,
            "data": request.data,
            "created_at": datetime.utcnow().isoformat()
        }

        # Add optional fields if provided
        if request.document_id:
            record["document_id"] = request.document_id
        if request.user_id:
            record["user_id"] = request.user_id

        # Insert into database
        response = supabase.table("data_extractions").insert(record).execute()

        if response.data:
            return SaveExtractionResponse(
                success=True,
                extraction_id=str(response.data[0]["id"]),
                message="Extraction saved successfully"
            )
        else:
            raise HTTPException(
                status_code=500,
                detail="Failed to save extraction - no data returned"
            )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to save extraction: {str(e)}"
        )


@router.get("/extractions")
async def list_extractions(
    user_id: Optional[str] = Query(default=None, description="Filter by user ID"),
    document_id: Optional[str] = Query(default=None, description="Filter by document ID"),
    extraction_type: Optional[str] = Query(default=None, description="Filter by type"),
    limit: int = Query(default=50, ge=1, le=100, description="Maximum results")
):
    """
    List saved extractions from the database.

    Supports filtering by user_id, document_id, and extraction_type.
    """
    try:
        query = supabase.table("data_extractions").select("*")

        if user_id:
            query = query.eq("user_id", user_id)
        if document_id:
            query = query.eq("document_id", document_id)
        if extraction_type:
            query = query.eq("extraction_type", extraction_type)

        query = query.order("created_at", desc=True).limit(limit)

        response = query.execute()

        return {
            "success": True,
            "count": len(response.data),
            "extractions": response.data
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to list extractions: {str(e)}"
        )


@router.get("/extraction/{extraction_id}")
async def get_extraction(extraction_id: str):
    """
    Get a specific extraction by ID.
    """
    try:
        response = supabase.table("data_extractions").select("*").eq("id", extraction_id).execute()

        if not response.data:
            raise HTTPException(
                status_code=404,
                detail="Extraction not found"
            )

        return {
            "success": True,
            "extraction": response.data[0]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to get extraction: {str(e)}"
        )


@router.delete("/extraction/{extraction_id}")
async def delete_extraction(extraction_id: str):
    """
    Delete an extraction by ID.
    """
    try:
        response = supabase.table("data_extractions").delete().eq("id", extraction_id).execute()

        return {
            "success": True,
            "message": "Extraction deleted successfully"
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to delete extraction: {str(e)}"
        )


# ============================================================================
# Legacy Endpoints (for backwards compatibility)
# ============================================================================

@router.post("/extract-tables")
async def extract_tables_legacy(
    document_id: str = Query(..., description="Document ID from database")
):
    """
    Legacy endpoint - Extract tables from a document stored in the system.

    Note: This endpoint requires the document to already be uploaded to Supabase storage.
    For direct PDF upload, use /upload-and-extract-tables instead.
    """
    try:
        # Get document info
        doc_response = supabase.table("documents").select("*").eq("id", document_id).execute()

        if not doc_response.data:
            raise HTTPException(
                status_code=404,
                detail="Document not found"
            )

        document = doc_response.data[0]
        storage_key = document.get("storage_key")

        if not storage_key:
            raise HTTPException(
                status_code=400,
                detail="Document has no associated file in storage"
            )

        # Download file from Supabase storage
        try:
            file_data = supabase.storage.from_("documents").download(storage_key)
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to download document from storage: {str(e)}"
            )

        # Process the PDF
        temp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                temp_file.write(file_data)
                temp_path = temp_file.name

            tables, total_pages = process_pdf_tables(temp_path)

            metadata = ExtractionMetadata(
                filename=document.get("filename", "unknown.pdf"),
                total_pages=total_pages,
                extraction_timestamp=datetime.utcnow().isoformat(),
                tables_found=len(tables),
                figures_found=0
            )

            return ExtractTablesResponse(
                success=True,
                metadata=metadata,
                tables=tables
            )

        finally:
            if temp_path and os.path.exists(temp_path):
                os.unlink(temp_path)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to extract tables: {str(e)}"
        )


@router.post("/extract-figures")
async def extract_figures_legacy(
    document_id: str = Query(..., description="Document ID from database"),
    include_image_data: bool = Query(default=False)
):
    """
    Legacy endpoint - Extract figures from a document stored in the system.

    Note: This endpoint requires the document to already be uploaded to Supabase storage.
    For direct PDF upload, use /upload-and-extract-figures instead.
    """
    try:
        doc_response = supabase.table("documents").select("*").eq("id", document_id).execute()

        if not doc_response.data:
            raise HTTPException(
                status_code=404,
                detail="Document not found"
            )

        document = doc_response.data[0]
        storage_key = document.get("storage_key")

        if not storage_key:
            raise HTTPException(
                status_code=400,
                detail="Document has no associated file in storage"
            )

        try:
            file_data = supabase.storage.from_("documents").download(storage_key)
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to download document from storage: {str(e)}"
            )

        temp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                temp_file.write(file_data)
                temp_path = temp_file.name

            figures, total_pages = process_pdf_figures(temp_path, include_image_data)

            metadata = ExtractionMetadata(
                filename=document.get("filename", "unknown.pdf"),
                total_pages=total_pages,
                extraction_timestamp=datetime.utcnow().isoformat(),
                tables_found=0,
                figures_found=len(figures)
            )

            return ExtractFiguresResponse(
                success=True,
                metadata=metadata,
                figures=figures
            )

        finally:
            if temp_path and os.path.exists(temp_path):
                os.unlink(temp_path)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to extract figures: {str(e)}"
        )
